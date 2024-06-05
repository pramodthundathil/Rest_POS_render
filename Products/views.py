from django.shortcuts import render, redirect, get_object_or_404
from .models import FoodCategory, Menu, Tables, Order, OrderItem, Checkout, Tax
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger 
from django.http import JsonResponse
from django.template.loader import render_to_string


from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from Home.decorators import admin_only, allowed_users

# Create your views here.


def Add_Category(request):
    if request.method == "POST":
        pic = request.FILES['pic']
        cname = request.POST['cname']
        foodcategory = FoodCategory.objects.create(image = pic, name= cname)
        foodcategory.save()
        messages.success(request,"Food Category Addedd...")
        return redirect("List_Category")
    
    return render(request,'add-category.html')

def List_Category(request):
    food_category = FoodCategory.objects.all()
    p = Paginator(food_category, 20)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
        "food_category":page_obj
    }
    return render(request,'list-category.html',context)

def DeleteCategory(request,pk):
    FoodCategory.objects.get(id = pk).delete()
    messages.success(request,'Food Category Deleted')
    return redirect('List_Category')


def AddTax(request):
    if request.method == "POST":
        name = request.POST.get('name')
        tax_rate = request.POST.get('tax')
        tax = Tax.objects.create(tax_name = name,tax_percentage = tax_rate )
        tax.save()
        messages.success(request,'Tax Value Added Success')
        return redirect("ListTax")
    return render(request,"add-tax-slab.html")

def ListTax(request):
    tax = Tax.objects.all()

    context = {
        "tax":tax
    }
    return render(request,"list-tax.html",context)


def Add_Product(request):
    food_category = FoodCategory.objects.all()
    tax = Tax.objects.all()
    description = " "
    if request.method == "POST":
        name = request.POST['name']
        category = FoodCategory.objects.get(id = int(request.POST['category']))
        potion = request.POST['potion']
        diet = request.POST['diet']
        price = request.POST['price']
        stock = request.POST['stock']
        image = request.FILES['pic']
        description = request.POST['description']
        tax_name = request.POST["tax_name"]
        tax_value = request.POST["tax_value"]

        menu = Menu.objects.create(
            name = name, 
            category = category, 
            image = image, 
            potion = potion, 
            diet =diet, 
            price = price, 
            stock = stock, 
            description = description,
            tax = tax_name,
            tax_value = Tax.objects.get(id = int(tax_value))

            )
        menu.save()
        messages.success(request,"Menu Item added Success...")
        return redirect("List_Product")
    
    context = {
        "food_category":food_category,
        "tax":tax

    }
    return render(request,'add-product.html',context)

def List_Product(request):
    menu = Menu.objects.all()

    context = {
        "menu":menu,
    }
    return render(request,'list-product.html',context)

def DeleteProduct(request,pk):
    menu  = Menu.objects.get(id = pk)
    if menu.status == False:
        menu.status = True
    else:
        menu.status = False
    menu.save()
    messages.info(request,"Product Deleted....")
    return redirect("List_Product")


def Add_Table(request):
    if request.method == "POST":
        tnum = request.POST['tnum']
        seats = request.POST['seats']
        if Tables.objects.filter(Table_number = tnum).exists():
            messages.error(request,"Table Already Exists...")
            return redirect("Add_Table")
        else:
            table = Tables.objects.create(Table_number = tnum, Number_of_Seats = seats)
            table.save()
            messages.success(request,"Table added Success...")
            return redirect("List_Table")

    return render(request,"add-table.html")

def List_Table(request):
    table = Tables.objects.all()
    context = {
        "table":table
    }
    return render(request,"list-table.html",context)

def Delete_Table(request,pk):
    Tables.objects.get(id = pk).delete()
    messages.success(request,"Table deleted success....")
    return redirect("List_Table")


def Pos(request):
    category = FoodCategory.objects.all()
    menu = Menu.objects.filter(status = True)
    table = Tables.objects.all()
    orders = Order.objects.filter(user = request.user, checkout_status = False)
    order_details = []

    for order in orders:
        total_items = sum(item.quantity for item in order.items.all())
        total_price = sum(item.get_total_price() for item in order.items.all())
        order_details.append({
            'order': order,
            'total_items': total_items,
            'total_price': total_price,
        })

    context = {
        "category":category,
        "menu":menu,
        "table":table,
        "orders":orders,
        'order_details': order_details,
    }
    return render(request,'posinterface.html',context)

@admin_only
def PosIndex(request):
    category = FoodCategory.objects.filter()
    menu = Menu.objects.filter(status = True)
    table = Tables.objects.all()
    orders = Order.objects.filter(checkout_status = False)
    order_details = []

    for order in orders:
        total_items = sum(item.quantity for item in order.items.all())
        total_price = sum(item.get_total_price() for item in order.items.all())
        order_details.append({
            'order': order,
            'total_items': total_items,
            'total_price': total_price,

        })

    context = {
        "category":category,
        "menu":menu,
        "table":table,
        "orders":orders,
        'order_details': order_details,
        
    }
    return render(request,'posinterface.html',context)

def OrderSingle(request,pk):
    category = FoodCategory.objects.all()
    menu = Menu.objects.filter(status = True)
    order = Order.objects.get(id = pk)
    item = OrderItem.objects.filter(order = order)
    total_price = sum(item.get_total_price() for item in order.items.all())
    


    context = {
        "category":category,
        "menu":menu,
        "order":order,
        "item":item,
        "total_price":round(total_price,2)
    }
    return render(request,"order-single.html",context)

def CreateOrder(request):
    if request.method == "POST":
        table = Tables.objects.get(id = int(request.POST['table']))
        order = Order.objects.create(table = table,user = request.user )
        order.save()
        
        return redirect('OrderSingle',pk = order.id )
    


def add_to_order(request):
    if request.method == "POST":
        menu_item_id = request.POST.get('menu_item_id')
        order_id = request.POST.get('order_id')
        menu_item = get_object_or_404(Menu, id=menu_item_id)
        order = get_object_or_404(Order, id=order_id)

        # Create or update the OrderItem
        order_item, created = OrderItem.objects.get_or_create(order=order, menu_item=menu_item, defaults={'price': menu_item.price})
        if not created:
            order_item.quantity += 1
            order_item.save()

        # Render the updated order items and return as HTML
        item = OrderItem.objects.filter(order = order)
        total_price = sum(item.get_total_price() for item in order.items.all())
        order_html = render_to_string('order-summery.html', {'order': order,"item":item,"total_price":total_price})

        return JsonResponse({'order_html': order_html})
    return JsonResponse({'error': 'Invalid request'}, status=400)

def increase_quantity(request):
    if request.method == "POST":
        item_id = request.POST.get('item_id')
        order_item = get_object_or_404(OrderItem, id=item_id)
        order_item.quantity += 1
        order_item.save()

        item = OrderItem.objects.filter(order = order_item.order)
        total_price = sum(item.get_total_price() for item in order_item.order.items.all())
        order_html = render_to_string('order-summery.html', {'order': order_item.order, 'total_price': total_price,"item":item})

        return JsonResponse({'order_html': order_html})
    return JsonResponse({'error': 'Invalid request'}, status=400)

def decrease_quantity(request):
    if request.method == "POST":
        item_id = request.POST.get('item_id')
        order_item = get_object_or_404(OrderItem, id=item_id)
        if order_item.quantity > 1:
            order_item.quantity -= 1
            order_item.save()

        item = OrderItem.objects.filter(order = order_item.order)
        total_price = sum(item.get_total_price() for item in order_item.order.items.all())
        order_html = render_to_string('order-summery.html', {'order': order_item.order, 'total_price': total_price,"item":item})

        return JsonResponse({'order_html': order_html})
    return JsonResponse({'error': 'Invalid request'}, status=400)


def Delete_menuitem(request,pk):
    item = OrderItem.objects.get(id = pk)
    order = item.order.id
    item.delete()
    return redirect("OrderSingle",pk = order)


def TakeOrder(request,pk):
    order = Order.objects.get(id =pk)
    order.take_order = True
    order.completion_status = False
    order.save()
    posted_data = order 
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        "updates",
        {
            "type": "send_update",
            "message": "Database updated",
            
        }
    )
    return redirect("Pos")


def receipt_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    items = order.items.all()
    total_price = sum(item.get_total_price() for item in items)
    context = {
        'order': order,
        'item': items,
        'total_price': round(total_price,2),
    }
    return render(request, 'receipt.html', context)


def KitchenDashboard(request):
    orders = Order.objects.filter( take_order= True, completion_status = False )
    order_details = []
    orderitem = OrderItem.objects.all()
    for order in orders:
        orderitem = OrderItem.objects.filter(order = order)
        order_details.append(
            {
                "order":order,
                "orderitem":orderitem
            }
        ) 

    context = {
        "order_details":order_details
    } 

    return render(request,"kitchendash.html",context)



def refresh_table(request):
    orders = Order.objects.filter(take_order = True,completion_status = False, checkout_status = False)
    order_details = []
    orderitem = OrderItem.objects.all()
    for order in orders:
        orderitem = OrderItem.objects.filter(order = order)
        order_details.append(
            {
                "order":order,
                "orderitem":orderitem
            }
        ) 

    context = {
        "order_details":order_details
    }
    table_html = render_to_string('kitchendashitems.html', context)
    return JsonResponse({'table_html': table_html})

def refresh_order(request):
    category = FoodCategory.objects.filter()
    menu = Menu.objects.filter(status = True)
    table = Tables.objects.all()
    orders = Order.objects.filter(checkout_status = False)
    order_details = []

    for order in orders:
        total_items = sum(item.quantity for item in order.items.all())
        total_price = sum(item.get_total_price() for item in order.items.all())
        order_details.append({
            'order': order,
            'total_items': total_items,
            'total_price': total_price,

        })

    context = {
        "category":category,
        "menu":menu,
        "table":table,
        "orders":orders,
        'order_details': order_details,
        
    }
    table_html = render_to_string('order-datas.html', context)
    return JsonResponse({'table_html': table_html})

def Status_Change(request):
    if request.method == "POST":
        order_id = request.POST.get('orderid')  # Corrected key name
        order = get_object_or_404(Order, id=order_id)
        order.status = "In Progress"
        order.save()
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
        "updates",
        {
            "type": "send_update",
            "message": "Database updated",
            
        }
        )
        
        orders = Order.objects.filter(take_order=True,completion_status = False, checkout_status = False)
        order_details = []
        for order in orders:
            orderitem = OrderItem.objects.filter(order=order)
            order_details.append({
                "order": order,
                "orderitem": orderitem
            }) 

        context = {
            "order_details": order_details
        }
        table_html = render_to_string('kitchendashitems.html', context)
        return JsonResponse({'order_html': table_html})
    
def Status_Change_Order_Ready(request):
    if request.method == "POST":
        order_id = request.POST.get('orderid')  # Corrected key name
        order = get_object_or_404(Order, id=order_id)
        order.status = "Order Ready"
        order.save()
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
        "updates",
        {
            "type": "send_update",
            "message": "Database updated",
            
        }
        )
        
        orders = Order.objects.filter(take_order=True,completion_status = False, checkout_status = False)
        order_details = []
        for order in orders:
            orderitem = OrderItem.objects.filter(order=order)
            order_details.append({
                "order": order,
                "orderitem": orderitem
            }) 

        context = {
            "order_details": order_details
        }
        table_html = render_to_string('kitchendashitems.html', context)
        return JsonResponse({'order_html': table_html})
    

def Status_Change_Menu_Finish(request):
    if request.method == "POST":
        order_item_id = request.POST.get('orderid')  # Ensure the correct key is used
        order_item = get_object_or_404(OrderItem, id=order_item_id)
        order_item.completion_status = True
        order_item.save()
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
        "updates",
        {
            "type": "send_update",
            "message": "Database updated",
            
        }
        )
        
        orders = Order.objects.filter(take_order=True,completion_status = False, checkout_status = False)
        order_details = []
        for order in orders:
            orderitem = OrderItem.objects.filter(order=order)
            order_details.append({
                "order": order,
                "orderitem": orderitem
            })

        context = {
            "order_details": order_details
        }
        table_html = render_to_string('kitchendashitems.html', context)
        return JsonResponse({'order_html': table_html})

    

def Status_Change_OrderCompeletion(request,pk):
    
    order = get_object_or_404(Order, id=pk)
    order.completion_status = True
    order.take_order = False
    order.save()
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
    "updates",
    {
        "type": "send_update",
        "message": "Database updated",
        
    }
    )
    return redirect("KitchenDashboard")

def calculate_tax(menu_item, quantity):
    tax_value = 0
    tax_value = menu_item.tax_amount * quantity
    return round(tax_value,2)

def SettleOrder(request, pk):
    order = get_object_or_404(Order, id=pk)
    if request.method == "POST":
        payment = request.POST.get("payment")
        order_items = OrderItem.objects.filter(order=order)
        total_price = sum(item.menu_item.price * item.quantity for item in order_items)
        total_tax_amount = sum(calculate_tax(item.menu_item, item.quantity) for item in order_items)

        checkout = Checkout.objects.create(
            order=order,
            payment_method=payment,
            payment_status="Paid",
            total_price=total_price,
            tax_amount=total_tax_amount
        )
        checkout.save()
        order.checkout_status = True
        order.save()
        messages.info(request, "Bill Settled....")
        return redirect("Pos")

    # Render the order settlement page if it's a GET request
    return render(request, 'settle_order.html', {'order': order})
    
@allowed_users(allowed_roles=["admin"])   
def ViewCheckouts(request):
    checkout = Checkout.objects.all()
    context = {
        "checkout":checkout
    }
    return render(request,"settledorders.html",context)


def Reports(request):
    return render(request,"reports.html")



from django.http import HttpResponse
from django.utils.timezone import now
import openpyxl
from openpyxl.styles import Font
from .models import Order, Checkout

def generate_excel_report(request):
    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Today\'s Orders Report'

    # Define the columns
    columns = ['Order ID', 'Order Date', 'Table', 'Total Price', 'Tax Amount', 'Payment Method', 'Payment Status']

    # Set the header row
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    # Fetch today's orders
    today = now().date()
    orders = Order.objects.filter(create_date__date=today)

    # Add the order data to the worksheet
    for order in orders:
        checkout = Checkout.objects.filter(order=order).first()
        row_num += 1
        worksheet.cell(row=row_num, column=1).value = order.id
        worksheet.cell(row=row_num, column=2).value = order.create_date.strftime('%Y-%m-%d %H:%M')
        worksheet.cell(row=row_num, column=3).value = order.table.Table_number if order.table else 'N/A'
        worksheet.cell(row=row_num, column=4).value = checkout.total_price if checkout else 'N/A'
        worksheet.cell(row=row_num, column=5).value = checkout.tax_amount if checkout else 'N/A'
        worksheet.cell(row=row_num, column=6).value = checkout.payment_method if checkout else 'N/A'
        worksheet.cell(row=row_num, column=7).value = checkout.payment_status if checkout else 'N/A'

    # Set the column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=todays_orders_report.xlsx'
    workbook.save(response)
    return response

def generate_orders_report(request):
    if request.method == "POST":
        # Get start date and end date from the request
        start_date = request.POST.get('sdate')
        end_date = request.POST.get('edate')

        # Create a new Excel workbook and add a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Orders Report'

        # Define the columns
        columns = ['Order ID', 'Order Date', 'Table', 'Total Price', 'Tax Amount', 'Payment Method', 'Payment Status']

        # Set the header row
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)

        # Fetch orders within the date range
        orders = Order.objects.filter(create_date__gte=start_date, create_date__lte=end_date)

        # Add the order data to the worksheet
        for order in orders:
            checkout = Checkout.objects.filter(order=order).first()
            row_num += 1
            worksheet.cell(row=row_num, column=1).value = order.id
            worksheet.cell(row=row_num, column=2).value = order.create_date.strftime('%Y-%m-%d %H:%M')
            worksheet.cell(row=row_num, column=3).value = order.table.Table_number if order.table else 'N/A'
            worksheet.cell(row=row_num, column=4).value = checkout.total_price if checkout else 'N/A'
            worksheet.cell(row=row_num, column=5).value = checkout.tax_amount if checkout else 'N/A'
            worksheet.cell(row=row_num, column=6).value = checkout.payment_method if checkout else 'N/A'
            worksheet.cell(row=row_num, column=7).value = checkout.payment_status if checkout else 'N/A'

        # Set the column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        # Create an HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=orders_report_{start_date}_to_{end_date}.xlsx'
        workbook.save(response)
        return response
    






        
        

    

